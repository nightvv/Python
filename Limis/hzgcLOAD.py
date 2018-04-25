#!/usr/bin/env python
# encoding: utf-8
"""
Author: ISeeMoon
Python: 3.6
Software: PyCharm
File: hzgc_load.py
Time: 2018/4/23 20:12
"""

from openpyxl import load_workbook

class hzgc_API(object):

    exce = load_workbook(r"C:\excelPend_8251.xlsx")
    sheet = exce.worksheets[0]
    sample_num = sheet.max_row

    def JPMC(self,i):
        return self.sheet['U'+str(i+1)].value #检品名称

    def PP(self,i):
        return self.sheet['AF'+str(i+1)].value #品牌

    def SSDL(self,i):
        return self.sheet['T'+str(i+1)].value #所属大类

    def SSXL(self,i):
        return '/' #所属小类

    def SCDW(self,i):
        if self.SFJK(i) == '进口':
            return '/'
        else:
            return self.sheet['AH'+str(i+1)].value #生产单位

    def CD(self,i):
        return self.sheet['AE'+str(i+1)].value + sheet['AG'+str(i+1)].value #产地

    def SFJK(self,i):
        return self.sheet['AC'+str(i+1)].value #是否进口

    def SCDWDZ(self,i):
        if self.SFJK(i) == '进口':
            return '/'
        else:
            return self. sheet['AJ'+str(i+1)].value #生产单位地址

    def JXSMC(self,i):
        if self.SFJK(i) == '进口':
            return self.sheet['AH'+str(i+1)].value #经销商名称
        else:
            return '/'

    def JXSDZ(self,i):
        if self.SFJK(i) == '进口':
            return self. sheet['AJ'+str(i+1)].value #经销商地址
        else:
            return '/'

    def JXSSF(self,i):
        if self.SFJK(i) == '进口':
            return self. sheet['AI'+str(i+1)].value #经销商省份
        else:
            return '/'

    def WTF(self,i):
        if len(self.sheet['AN'+str(i+1)].value) == 0:
            return '/'
        else:
            return self.sheet['AN'+str(i+1)].value #委托方

    def SSDQ(self,i):
        return self.sheet['AO'+str(i+1)].value.split('省')[0] #所属地区

    def SZD(self,i):
        return self.sheet['AO'+str(i+1)].value.split('省')[-1].split('市')[0]

    def WTFDZ(self,i):
        if len(self.sheet['AO'+str(i+1)].value) == 0:
            return '/'
        else:
            return self.sheet['AO'+str(i+1)].value #委托方地址

    def GYDW(self,i):
        return self.sheet['AP'+str(i+1)].value #供样单位

    def GYDWXZ(self,i):
        return '其它' #供样单位性质

    def GYDWGSD(self,i):
        return '杭州' #供样单位归属地

    def GYDWDZ(self,i):
        return self.sheet['AQ'+str(i+1)].value #供样单位地址

    def GYDWLXR(self,i):
        return self.sheet['AS'+str(i+1)].value #供样单位联系人

    def GYDWLXDH(self,i):
        return self.sheet['AT'+str(i+1)].value #供样单位联系电话

    def GYDWYX(self,i):
        return '/' #供样单位邮箱

    def BCYDW(self,i):
        return self.sheet['G'+str(i+1)].value #被抽样单位

    def BCYDWXZ(self,i):
        return self.sheet['S'+str(i+1)].value #被抽样单位性质

    def BCYDWSZD(self,i):
        return self.sheet['I'+str(i+1)].value #被抽样单位所在地

    def DYLX(self,i):
        return self.sheet['L'+str(i+1)].value #地域类型

    def BCYDWSF(self,i):
        return self.sheet['H'+str(i+1)].value #被抽样单位省份

    def BCYDWGSD(self,i):
        return self.BCYDWSZD(i) #被抽样单位归属地

    def BCYDWXM(self,i):
        return self.sheet['J'+str(i+1)].value #被抽样单位县名

    def BCYDWDZ(self,i):
        return self.sheet['K'+str(i+1)].value #被抽样单位地址

    def BCYDWFZR(self,i):
        return self.sheet['O'+str(i+1)].value #被抽样单位负责人

    def BCYDWLXDH(self,i):
        return self.sheet['P'+str(i+1)].value #被抽样单位联系电话

    def CYJS(self,i):
        return self.sheet['AD'+str(i+1)].value #抽样基数

    def DJ(self,i):
        return self.sheet['AV'+str(i+1)].value.spli1t('单价：')[-1].split(' ')[0] #单价

    def ZXBZ(self,i):
        return '/' #执行标准

    def GG(self,i):
        return self.sheet['AA'+str(i+1)].value #规格

    def BZ(self,i):
        if '盒' in self.CYJS(i):
            return '纸盒'
        else:
            return '塑瓶'
    def BZGG(self,i):
        return self.sheet['AA'+str(i+1)].value #包装规格
    def JX(self,i):
        return '/' #剂型

    def JYYJ(self,i):
        return self.sheet['AW'+str(i+1)].value #检验依据

    def JYMDDX(self,i):
        return '国家级抽验' #检验目的大项

    def JYMDXX(self,i):
        return '2018年国家化妆品监督抽检' #检验目的小项

    def JPSL(self,i):
        return int(self.sheet['Z'+str(i+1)].value)-1 #检品数量

    def JPSLDW(self,i):
        return self.sheet['AD'+str(i+1)].value[-1] #检品数量单位

    def LYSL(self,i):
        return self.sheet['AV'+str(i+1)].value.split('留样：')[-1][:-1].split(self.JPSLDW(i))[0] #留样数量

    def PZWH(self,i):
        return self.sheet['Y'+str(i+1)].value #批准文号

    def RWLY(self,i):
        return '食药监药化监（2018）17号' #任务来源

    def BSFL(self,i):
        return '2018年中央补助地方任务' #报送分类

    def PH(self,i):
        return self.sheet['V'+str(i+1)].value #批号

    def SCRQ(self,i):
        return self.sheet['W'+str(i+1)].value #生产日期

    def BZQ(self,i):
        return self.sheet['X'+str(i+1)].value #保质期

    def YSHWT(self,i):
        return '/' #颜色和物态

    def CYDBH(self,i):
        return self.sheet['E'+str(i+1)].value #抽样单编号

    def XKZH(self,i):
        return self.sheet['AK'+str(i+1)].value #许可证号

    def ZJKS(self,i):
        return '保化室' #主检科室

    def JYXM(self,i):
        return '部分检验' #检验项目

    def JYZQ(self,i):
        if 'GC' in self.CYDBH(i):
            return '20' #检验周期
        else:
            return '25'

    def CYRY(self,i):
        return '翁文涛 费文翔' #抽样人员

    def CYSJ(self,i):
        return self.sheet['F'+str(i+1)].value #抽样时间


s = hzgc_API()
s.SCDW(3)
s.JXSMC(3)