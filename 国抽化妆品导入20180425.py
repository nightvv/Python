#!/usr/bin/env python
# encoding: utf-8 
"""
Author: ISeeMoon
Python: 3.6
Software: PyCharm
File: 1234.py
Time: 2018/4/22 22:08
"""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.select import Select
from prettytable import PrettyTable
from openpyxl import load_workbook

print('本软件目前仅适用于【化妆品国抽】导入。\n【委托方所属地区】、【样品包装方式】和【抽样人员】需要自行核对！！！\n【委托方所属地区】、【样品包装方式】和【抽样人员】需要自行核对！！！\n【委托方所属地区】、【样品包装方式】和【抽样人员】需要自行核对！！！\n -------2018-04-23 Version 1.0-------')

exce = load_workbook(r"C:\excelPend_8251.xlsx")
sheet = exce.worksheets[0]
sample_num = sheet.max_row

limis_name = input('请输入Limis用户名(按ENTER确认):')
limis_pwd = input('请输入Limis密码(按ENTER确认):')

def login():
    global wb
    wb = webdriver.Chrome("C:\chromedriver.exe")
    wb.implicitly_wait(30)
    wb.maximize_window()
    wb.get('http://10.0.144.50/web/Login.aspx')
    wb.find_element_by_xpath("//input[@id='txtUserCode']").send_keys(limis_name)
    wb.find_element_by_xpath("//input[@id='txtUserPassword']").send_keys(limis_pwd)
    wb.find_element_by_xpath("//a[@id='btnLogin']").click()
    time.sleep(2)
    
def selecttype():
    category = {'1':'食品','2':'药品','3':'保健品','4':'化妆品','5':'药包材'}
    category_table = PrettyTable(['类别编号','类别名称'])
    category_table.add_row(['1','食品'])
    category_table.add_row(['2', '药品'])
    category_table.add_row(['3', '保健品'])
    category_table.add_row(['4', '化妆品'])
    category_table.add_row(['5', '药包材'])
    print(category_table)
    index_category = input('请输入检品类别编号(按ENTER确认):')
    if index_category == str(1):
        pass
    elif index_category == str(2):
        pass
    elif index_category == str(3):
        pass
    elif index_category == str(4):
        print('\n\n已选定检品类别为→化妆品\n\n')
        HZ_type_table = PrettyTable(['小类编号','小类名称'])
        HZ_type_table.add_row(['1','进口非特'])
        HZ_type_table.add_row(['2', '国产特殊'])
        HZ_type_table.add_row(['3', '抽验检验'])
        print(HZ_type_table)
    else:
        pass
    type_index = input('请选择检品检验小类(按ENTER确认):')
    type_dic = {'食品':[],
            '药品':[],
            '保健品': [],
            '化妆品': {'1':'5','2':'6','3':'2'},
            '药包材': [],}
    index_1 = {'1':'1','2':'2','3':'5','4':'3','5':'4'}[index_category]
    index_2 = type_dic[category[index_category]][type_index]
    wb.find_element_by_xpath("//ul[@id='leftMenuTree']/li[1]/ul/li[1]/ul/li[{}]/ul/li[{}]".format(index_1,index_2)).click()
    time.sleep(2)

def save_sample():
    wb.switch_to.frame(wb.find_element_by_xpath("//iframe"))    
    #添加按钮
    wb.find_element_by_xpath("//div[@class='datagrid-toolbar']//span[contains(text(),'添加')]").click()

    wb.switch_to.frame(wb.find_element_by_xpath("//iframe[@id='ifm_first']"))
    wb.switch_to.frame(wb.find_element_by_xpath("//iframe[@id='ifm_sample']"))

    #检品名称
    wb.find_element_by_xpath("//div[@id='p-first']/ul[2]/li[2]/input").send_keys(JPMC)
    #品牌PP
    wb.find_element_by_xpath("//div[@id='p-first']/ul[2]/li[5]/input").send_keys(PP)
    #所属大类SSDL
    if SSDL in '染发类产品':
        Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[3]/li[2]/select")).select_by_value('染发类产品')
    elif SSDL in '防晒类产品':
        Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[3]/li[2]/select")).select_by_value('防晒类产品')
    elif SSDL in '祛痘产品':
        Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[3]/li[2]/select")).select_by_value('祛痘产品')    
    #所属小类SSXL
    wb.find_element_by_xpath("//div[@id='p-first']/ul[3]/li[5]/input").send_keys(SSXL)
    #生产单位SCDW
    wb.find_element_by_xpath("//div[@id='p-first']/ul[4]/li[2]/span/input[2]").send_keys(SCDW)
    #产地CD
    wb.find_element_by_xpath("//div[@id='p-first']/ul[4]/li[5]/input").send_keys(CD)
    #是否进口SFJK
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[4]/li[8]/select")).select_by_value(SFJK)
    #生产单位地址SCDWDZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[5]/li[2]/input").send_keys(SCDWDZ)
    #经销商名称
    wb.find_element_by_xpath("//div[@id='p-first']/ul[6]/li[2]/input").send_keys(JXSMC)
    #经销商省份JXSSF
    wb.find_element_by_xpath("//div[@id='p-first']/ul[6]/li[5]/input").send_keys(JXSSF)
    #经销商地址JXSDZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[7]/li[2]/input").send_keys(JXSDZ)
    #委托方WTF
    wb.find_element_by_xpath("//div[@id='p-first']/ul[8]/li[2]/span/input[2]").send_keys(WTF)
    #所属地区SSDQ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[8]/li[5]/input").send_keys(SSDQ)
    #所在地SZD
    wb.find_element_by_xpath("//div[@id='p-first']/ul[8]/li[8]/input").send_keys(SZD)
    #委托方地址WTFDZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[9]/li[2]/input").send_keys(WTFDZ)
    #供养单位GYDW
    wb.find_element_by_xpath("//div[@id='p-first']/ul[10]/li[2]/span/input[2]").send_keys(GYDW)
    #供养单位性质GYDWXZ
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[10]/li[5]/select")).select_by_value(GYDWXZ)
    #供养单位归属地GYDWGSD
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[10]/li[8]/select")).select_by_value(GYDWGSD)
    #供养单位地址GYDWDZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[11]/li[2]/input").send_keys(GYDWDZ)
    #供养单位联系人GYDWLXR
    wb.find_element_by_xpath("//div[@id='p-first']/ul[12]/li[2]/input").send_keys(GYDWLXR)
    #供养单位联系电话GYDWLXDH
    wb.find_element_by_xpath("//div[@id='p-first']/ul[12]/li[5]/input").send_keys(GYDWLXDH)
    #供养单位邮箱GYDWYX
    wb.find_element_by_xpath("//div[@id='p-first']/ul[12]/li[8]/input").send_keys(GYDWYX)
    #被抽样单位BCYDW
    wb.find_element_by_xpath("//div[@id='p-first']/ul[13]/li[2]/span/input[2]").send_keys(BCYDW)
    #被抽样单位性质BCYDWXZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[13]/li[5]/input").send_keys(BCYDWXZ)
    #被抽样单位所在地
    wb.find_element_by_xpath("//div[@id='p-first']/ul[14]/li[2]/input").send_keys(BCYDWSZD)
    #地域类型
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[14]/li[5]/select")).select_by_value(DYLX)
    #被抽样单位省份BCYDWSF
    wb.find_element_by_xpath("//div[@id='p-first']/ul[15]/li[2]/input").send_keys(BCYDWSF)
    #被抽样单位归属地BCYDWGSD
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[15]/li[5]/select")).select_by_value(BCYDWGSD[:-1])
    #被抽样单位县名BCYDWXM
    wb.find_element_by_xpath("//div[@id='p-first']/ul[15]/li[8]/input").send_keys(BCYDWXM)
    #被抽样单位地址
    wb.find_element_by_xpath("//div[@id='p-first']/ul[16]/li[2]/input").send_keys(BCYDWDZ)
    #被抽样单位负责人
    wb.find_element_by_xpath("//div[@id='p-first']/ul[17]/li[2]/input").send_keys(BCYDWFZR)
    #被抽样单位联系电话BCYDWLXDH
    wb.find_element_by_xpath("//div[@id='p-first']/ul[17]/li[5]/input").send_keys(BCYDWLXDH)
    #抽样基数
    wb.find_element_by_xpath("//div[@id='p-first']/ul[18]/li[2]/input").send_keys(CYJS)
    #单价DJ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[18]/li[5]/input").send_keys(DJ)
    #执行标准ZXBZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[18]/li[8]/input").send_keys(ZXBZ)
    #规格GG
    wb.find_element_by_xpath("//div[@id='p-first']/ul[19]/li[2]/span/input[2]").send_keys(GG)
    #包装BZ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[19]/li[5]/span/input[2]").send_keys(BZ)
    #包装规格BZGG
    wb.find_element_by_xpath("//div[@id='p-first']/ul[20]/li[2]/span/input[2]").send_keys(BZGG)
    #剂型JX
    wb.find_element_by_xpath("//div[@id='p-first']/ul[20]/li[5]/span/input[2]").send_keys(JX)
    #检验依据JYYJ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[21]/li[2]/span/input[2]").send_keys(JYYJ)
    #检验目的大项    
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[22]/li[2]/select")).select_by_value(JYMDDX )
    #检验目的小项
    wb.find_element_by_xpath("//div[@id='p-first']/ul[22]/li[5]/input").send_keys(JYMDXX)
    #批数
    PS = '1'
    wb.find_element_by_xpath("//div[@id='p-first']/ul[23]/li[2]/input").send_keys(PS)
    #检品数量JPSL
    wb.find_element_by_xpath("//div[@id='p-first']/ul[23]/li[5]/input").send_keys(JPSL)
    #检品数量单位JPSLDW
    wb.find_element_by_xpath("//div[@id='p-first']/ul[23]/li[8]/input").send_keys(JPSLDW)
    #留样数量LYSL
    wb.find_element_by_xpath("//div[@id='p-first']/ul[23]/li[11]/input").send_keys(LYSL)
    #批准文号PZWH
    wb.find_element_by_xpath("//div[@id='p-first']/ul[24]/li[2]/input").send_keys(PZWH)
    #任务来源RWLY
    wb.find_element_by_xpath("//div[@id='p-first']/ul[24]/li[5]/input").send_keys(RWLY)
    #报送分类BSFL
    wb.find_element_by_xpath("//div[@id='p-first']/ul[24]/li[8]/input").send_keys(BSFL)
    #批号PH
    wb.find_element_by_xpath("//div[@id='p-first']/ul[25]/li[2]/input").send_keys(PH)
    #生产日期SCRQ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[25]/li[5]/input").send_keys(SCRQ)
    #保质期BZQ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[25]/li[8]/input").send_keys(BZQ)
    #颜色和物态YSHWT
    wb.find_element_by_xpath("//div[@id='p-first']/ul[26]/li[2]/input").send_keys(YSHWT)
    #抽样单编号CYDBH
    wb.find_element_by_xpath("//div[@id='p-first']/ul[26]/li[5]/input").send_keys(CYDBH)
    #许可证号XKZH
    wb.find_element_by_xpath("//div[@id='p-first']/ul[27]/li[2]/input").send_keys(XKZH)
    #主检科室ZJKS
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[28]/li[2]/select")).select_by_value(ZJKS)
    #检验项目JYXM
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[28]/li[5]/select")).select_by_value(JYXM)
    #检验周期JYZQ
    Select(wb.find_element_by_xpath("//div[@id='p-first']/ul[28]/li[8]/select")).select_by_value(JYZQ)
    #抽样人员CYRY
    wb.find_element_by_xpath("//div[@id='p-first']/ul[31]/li[2]/input").send_keys(CYRY)
    #抽样时间CYSJ
    wb.find_element_by_xpath("//div[@id='p-first']/ul[31]/li[5]/input").send_keys(CYSJ)
    #保存检品信息
    wb.find_element_by_xpath("//div[@id='active_area']/div").click()
    time.sleep(3)
    wb.switch_to.default_content()

login()
selecttype()
for i in range(2,sample_num+1):
    JPMC = sheet['U'+str(i)].value #检品名称
    PP = sheet['AF'+str(i)].value #品牌
    SSDL = sheet['T'+str(i)].value #所属大类
    SSXL = '/' #所属小类
    SCDW = sheet['AH'+str(i)].value #生产单位
    CD = sheet['AE'+str(i)].value + sheet['AG'+str(i)].value #产地
    SFJK = sheet['AC'+str(i)].value #是否进口
    SCDWDZ = sheet['AJ'+str(i)].value #生产单位地址
    if SFJK == '进口':
        JXSMC = SCDW
        JXSDZ = SCDWDZ
        JXSSF = sheet['AI'+str(i)].value
        SCDW = '/'
        SCDWDZ = '/'
    elif SFJK == '国产':
        JXSMC = '/'
        JXSDZ = '/'
        JXSSF = '/'
    WTF = sheet['AN'+str(i)].value #委托方
    SSDQ = sheet['AO'+str(i)].value.split('省')[0] #所属地区
    SZD = sheet['AO'+str(i)].value.split('省')[-1].split('市')[0]
    WTFDZ = sheet['AO'+str(i)].value #委托方地址
    if len(WTF) == 0 or len(WTFDZ) == 0:
        WTF = '/'
        WTFDZ = '/'
    GYDW = sheet['AP'+str(i)].value #供样单位
    GYDWXZ = '其它' #供样单位性质
    GYDWGSD = '杭州' #供样单位归属地
    GYDWDZ = sheet['AQ'+str(i)].value #供样单位地址
    GYDWLXR = sheet['AS'+str(i)].value #供样单位联系人
    GYDWLXDH = sheet['AT'+str(i)].value #供样单位联系电话
    GYDWYX = '/' #供样单位邮箱
    BCYDW = sheet['G'+str(i)].value #被抽样单位
    BCYDWXZ = sheet['S'+str(i)].value #被抽样单位性质
    BCYDWSZD = sheet['I'+str(i)].value #被抽样单位所在地
    DYLX = sheet['L'+str(i)].value #地域类型
    BCYDWSF = sheet['H'+str(i)].value #被抽样单位省份
    BCYDWGSD = sheet['I'+str(i)].value #被抽样单位归属地
    BCYDWXM = sheet['J'+str(i)].value #被抽样单位县名
    BCYDWDZ = sheet['K'+str(i)].value #被抽样单位地址
    BCYDWFZR = sheet['O'+str(i)].value #被抽样单位负责人
    BCYDWLXDH = sheet['P'+str(i)].value #被抽样单位联系电话
    CYJS = sheet['AD'+str(i)].value #抽样基数
    DJ = sheet['AV'+str(i)].value.split('单价：')[-1].split(' ')[0] #单价
    ZXBZ = '/' #执行标准
    GG = sheet['AA'+str(i)].value #规格
    if '盒' in CYJS:
        BZ = '纸盒'
    else:
        BZ = '塑瓶'
    BZGG = sheet['AA'+str(i)].value #包装规格
    JX = '/' #剂型
    JYYJ = sheet['AW'+str(i)].value #检验依据
    JYMDDX = '国家级抽验' #检验目的大项
    JYMDXX = '2018年国家化妆品监督抽检' #检验目的小项
    JPSL = int(sheet['Z'+str(i)].value)-1 #检品数量
    JPSLDW = sheet['AD'+str(i)].value[-1] #检品数量单位
    LYSL = sheet['AV'+str(i)].value.split('留样：')[-1][:-1].split(JPSLDW)[0] #留样数量
    PZWH = sheet['Y'+str(i)].value #批准文号
    RWLY = '食药监药化监（2018）17号' #任务来源
    BSFL = '2018年中央补助地方任务' #报送分类
    PH = sheet['V'+str(i)].value #批号
    SCRQ = sheet['W'+str(i)].value #生产日期
    BZQ = sheet['X'+str(i)].value #保质期
    YSHWT = '/' #颜色和物态
    CYDBH = sheet['E'+str(i)].value #抽样单编号
    XKZH = sheet['AK'+str(i)].value #许可证号
    ZJKS = '保化室' #主检科室
    JYXM = '部分检验' #检验项目
    if 'GC' in CYDBH:
        JYZQ = '20' #检验周期
    elif 'SC' in CYDBH:
        JYZQ = '25'
    CYRY = '翁文涛 费文翔' #抽样人员
    CYSJ = sheet['F'+str(i)].value #抽样时间
    save_sample()
    print('已成功导入第{}个化妆品，共{}个化妆品'.format(i-1,sample_num-1))